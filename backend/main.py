from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Any
import pandas as pd
import io
import pdfplumber
import re
import pytesseract
from PIL import Image
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
def root():
    return {"message": "Backend çalışıyor"}


# -----------------------------
# Yardımcı Fonksiyonlar
# -----------------------------
def safe_str(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def safe_float(value: Any) -> float:
    text = safe_str(value).replace(",", ".")
    filtered = ""

    for ch in text:
        if ch.isdigit() or ch in [".", "-"]:
            filtered += ch

    try:
        return float(filtered) if filtered else 0.0
    except:
        return 0.0


def find_column(df_columns, keywords):
    for col in df_columns:
        lower = str(col).lower()
        for keyword in keywords:
            if keyword in lower:
                return col
    return None


def parse_offer_row(row, col_map, fallback_firma=""):
    return {
        "urunKodu": safe_str(row.get(col_map.get("urunKodu"), "")),
        "urunAciklamasi": safe_str(row.get(col_map.get("urunAciklamasi"), "")),
        "talepEdilenAdet": safe_float(row.get(col_map.get("talepEdilenAdet"), 0)),
        "firmaAdi": safe_str(row.get(col_map.get("firmaAdi"), fallback_firma)),
        "firmaAdedi": safe_float(row.get(col_map.get("firmaAdedi"), 0)),
        "paraBirimi": safe_str(row.get(col_map.get("paraBirimi"), "TRY")) or "TRY",
        "birimFiyat": safe_float(row.get(col_map.get("birimFiyat"), 0)),
        "iskonto": safe_float(row.get(col_map.get("iskonto"), 0)),
        "vade": safe_str(row.get(col_map.get("vade"), "")),
        "termin": safe_str(row.get(col_map.get("termin"), "")),
    }


def get_net_birim_fiyat(row, exchange_rates):
    para_birimi = safe_str(row.get("paraBirimi", "TRY")).upper() or "TRY"
    kur = exchange_rates.get(para_birimi, 1) or 1
    birim_fiyat = safe_float(row.get("birimFiyat", 0))
    iskonto = safe_float(row.get("iskonto", 0))

    net_fiyat = birim_fiyat * (1 - iskonto / 100)
    return net_fiyat * kur


def parse_vade_to_days(vade_text: str) -> int:
    text = safe_str(vade_text).lower()
    match = re.search(r"(\d+)", text)
    if not match:
        return 0
    return int(match.group(1))


def parse_termin_score(termin_text: str) -> int:
    text = safe_str(termin_text).lower()

    if "stok" in text:
        return 0
    if "aynı gün" in text or "ayni gün" in text:
        return 0
    if "1 gün" in text or "1 gun" in text:
        return 1
    if "hafta" in text:
        match = re.search(r"(\d+)", text)
        if match:
            return int(match.group(1)) * 7
        return 7
    if "ay" in text:
        match = re.search(r"(\d+)", text)
        if match:
            return int(match.group(1)) * 30
        return 30

    match = re.search(r"(\d+)", text)
    if match:
        return int(match.group(1))

    return 999


def extract_text_from_image_bytes(contents: bytes) -> str:
    image = Image.open(io.BytesIO(contents))
    text = pytesseract.image_to_string(image, lang="tur+eng")
    return text


def parse_offer_text_lines(text: str, fallback_firma=""):
    rows = []

    for raw_line in text.split("\n"):
        line = safe_str(raw_line)
        if not line:
            continue

        if len(line) < 3:
            continue

        rows.append({
            "urunKodu": "",
            "urunAciklamasi": line,
            "talepEdilenAdet": 0,
            "firmaAdi": fallback_firma,
            "firmaAdedi": 0,
            "paraBirimi": "TRY",
            "birimFiyat": 0,
            "iskonto": 0,
            "vade": "",
            "termin": "",
        })

    return rows


# -----------------------------
# Talepler / eski endpointler
# -----------------------------
@app.post("/parse-excel")
async def parse_excel(file: UploadFile = File(...)):
    contents = await file.read()
    df = pd.read_excel(io.BytesIO(contents))

    df.columns = [str(col).lower() for col in df.columns]

    urun = None
    miktar = None
    birim = None

    for col in df.columns:
        if "ürün" in col or "urun" in col or "aciklama" in col:
            urun = col
        if "miktar" in col or "adet" in col:
            miktar = col
        if "birim" in col:
            birim = col

    if not urun or not miktar:
        return {"error": "Gerekli kolonlar bulunamadı"}

    result = []
    for _, row in df.iterrows():
        result.append({
            "urun": str(row[urun]),
            "miktar": float(row[miktar]) if str(row[miktar]).strip() else 0,
            "birim": str(row[birim]) if birim else ""
        })

    return {"data": result}


@app.post("/parse-pdf")
async def parse_pdf(file: UploadFile = File(...)):
    contents = await file.read()
    rows = []

    units = [
        "adet", "ad", "kg", "lt", "l", "metre", "mt",
        "paket", "kutu", "torba", "takım", "takim", "plaka", "tüp", "tup"
    ]

    with pdfplumber.open(io.BytesIO(contents)) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            if not text:
                continue

            lines = text.split("\n")
            for line_index, line in enumerate(lines, start=1):
                cleaned = str(line).strip()
                if not cleaned:
                    continue

                lower = cleaned.lower()

                if (
                    "s.no" in lower
                    or "sıra no" in lower
                    or "aciklama" in lower
                    or "açıklama" in lower
                    or "birim" in lower
                    or "miktar" in lower
                ):
                    continue

                parts = cleaned.split()
                if len(parts) < 3:
                    continue

                if parts[0].isdigit():
                    parts = parts[1:]

                if parts and parts[0].isdigit():
                    parts = parts[1:]

                miktar = ""
                birim = ""
                urun_parts = parts[:]

                if urun_parts:
                    last = urun_parts[-1].replace(",", ".")
                    try:
                        float(last)
                        miktar = urun_parts.pop(-1)
                    except:
                        pass

                if urun_parts:
                    last_lower = urun_parts[-1].lower()
                    if last_lower in units:
                        birim = urun_parts.pop(-1)

                urun = " ".join(urun_parts).strip()

                if not urun:
                    continue

                rows.append({
                    "sayfa": page_index,
                    "satirNo": line_index,
                    "urun": urun,
                    "birim": birim,
                    "miktar": miktar,
                    "hamMetin": cleaned
                })

    return {
        "id": f"{file.filename}-pdf-0",
        "fileName": file.filename,
        "sourceType": "pdf",
        "label": f"PDF - {file.filename}",
        "rows": rows,
        "columns": ["sayfa", "satirNo", "urun", "birim", "miktar", "hamMetin"]
    }


# -----------------------------
# Teklif Analizi
# -----------------------------
@app.post("/parse-offer-file")
async def parse_offer_file(file: UploadFile = File(...)):
    filename = safe_str(file.filename)
    lower_name = filename.lower()
    contents = await file.read()

    # EXCEL
    if lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
       
        df = pd.read_excel(io.BytesIO(contents), header=3)
        df.columns = [safe_str(col) for col in df.columns]
        print("DF SHAPE:", df.shape)
        print("DF HEAD:")
        print(df.head(12).to_dict(orient="records"))
        print("EXCEL KOLONLARI:", list(df.columns))
        print("EXCEL DOSYA:", filename)
        

        col_map = {
            "urunKodu": find_column(df.columns, ["ürün kod", "urun kod", "stok kod", "malzeme kod", "kod"]),
            "urunAciklamasi": find_column(df.columns, ["ürün açıklaması", "urun aciklamasi", "ürün adı", "urun adi", "açıklama", "aciklama", "malzeme", "kalem", "description"]),
            "talepEdilenAdet": find_column(df.columns, ["talep edilen", "istenen adet", "talep adet", "talep miktar", "istenen", "ihtiyaç"]),
            "firmaAdi": find_column(df.columns, ["firma", "tedarikçi", "tedarikci", "satıcı", "satici"]),
            "firmaAdedi": find_column(df.columns, ["firma adedi", "verilen adet", "teklif adet", "adet", "miktar", "qty", "quantity"]),
            "paraBirimi": find_column(df.columns, ["para birimi", "pb", "döviz", "doviz", "currency", "curr"]),
            "birimFiyat": find_column(df.columns, ["birim fiyat", "fiyat", "unit price", "price"]),
            "iskonto": find_column(df.columns, ["iskonto", "indirim", "discount"]),
            "vade": find_column(df.columns, ["vade", "ödeme", "odeme", "payment"]),
            "termin": find_column(df.columns, ["termin", "teslim", "lead time", "delivery"]),
        }

        print("EXCEL COLUMN MAP:", col_map)

        result_rows = []
        seen = set()

        for _, row in df.iterrows():
            parsed = parse_offer_row(row, col_map)
            print("LOOP SATIRI:", safe_str(row.get("ürün kodu", "")), safe_str(row.get("ürün açıklaması", "")))
            if not parsed["firmaAdi"]:
                parsed["firmaAdi"] = filename.rsplit(".", 1)[0]
            row_has_any_value = any(safe_str(v) for v in row.tolist())
            if not row_has_any_value:
                continue
            if (
                not parsed["urunAciklamasi"]
                and not parsed["urunKodu"]
                and parsed["firmaAdedi"] == 0
                and parsed["birimFiyat"] == 0
             ):
                continue
            key = (
                safe_str(parsed["urunKodu"]),
                safe_str(parsed["urunAciklamasi"]),
                safe_str(parsed["firmaAdi"]),
                safe_float(parsed["firmaAdedi"]),
                safe_float(parsed["birimFiyat"]),
                safe_str(parsed["vade"]),
                safe_str(parsed["termin"]),
            )       
            if key in seen:
                continue
            seen.add(key)
            result_rows.append(parsed)
            
            print("APPEND EDILDI:", parsed["urunKodu"], parsed["urunAciklamasi"])
            print("TOPLAM:", len(result_rows))
            print("EXCEL PARSED ROW COUNT:", len(result_rows))
            print("EXCEL FIRST 3 ROWS:", result_rows[:3])

        return {
            "id": f"{filename}-excel",
            "fileName": filename,
            "sourceType": "excel",
            "label": f"Excel - {filename}",
            "columns": [
                "urunKodu",
                "urunAciklamasi",
                "talepEdilenAdet",
                "firmaAdi",
                "firmaAdedi",
                "paraBirimi",
                "birimFiyat",
                "iskonto",
                "vade",
                "termin",
            ],
            "rows": result_rows,
        }


    # PDF
    if lower_name.endswith(".pdf"):
        rows = []

        with pdfplumber.open(io.BytesIO(contents)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()

                if table and len(table) > 1:
                    headers = [safe_str(h) for h in table[0]]
                    df = pd.DataFrame(table[1:], columns=headers)

                    col_map = {
                        "urunKodu": find_column(df.columns, ["ürün kod", "urun kod", "stok kod", "kod"]),
                        "urunAciklamasi": find_column(df.columns, ["ürün", "urun", "açıklama", "aciklama", "malzeme"]),
                        "talepEdilenAdet": find_column(df.columns, ["talep edilen", "istenen adet", "talep adet", "talep miktar"]),
                        "firmaAdi": find_column(df.columns, ["firma", "tedarikçi", "tedarikci", "satıcı", "satici"]),
                        "firmaAdedi": find_column(df.columns, ["firma adedi", "verilen adet", "teklif adet", "adet", "miktar"]),
                        "paraBirimi": find_column(df.columns, ["para birimi", "pb", "döviz", "doviz", "currency"]),
                        "birimFiyat": find_column(df.columns, ["birim fiyat", "fiyat", "unit price"]),
                        "iskonto": find_column(df.columns, ["iskonto", "indirim"]),
                        "vade": find_column(df.columns, ["vade", "ödeme", "odeme"]),
                        "termin": find_column(df.columns, ["termin", "teslim", "lead time"]),
                    }

                    for _, row in df.iterrows():
                        parsed = parse_offer_row(row, col_map, fallback_firma=filename.rsplit(".", 1)[0])

                        if not parsed["urunAciklamasi"] and not parsed["urunKodu"]:
                            continue

                        rows.append(parsed)

        return {
            "id": f"{filename}-pdf",
            "fileName": filename,
            "sourceType": "pdf",
            "label": f"PDF - {filename}",
            "columns": [
                "urunKodu",
                "urunAciklamasi",
                "talepEdilenAdet",
                "firmaAdi",
                "firmaAdedi",
                "paraBirimi",
                "birimFiyat",
                "iskonto",
                "vade",
                "termin",
            ],
            "rows": rows,
        }

    # IMAGE
    if lower_name.endswith(".png") or lower_name.endswith(".jpg") or lower_name.endswith(".jpeg"):
        print("GÖRSEL DOSYA GELDİ:", filename)

        text = extract_text_from_image_bytes(contents)

        print("OCR TEXT BASLADI")
        print(text)
        print("OCR TEXT BITTI")

        rows = parse_offer_text_lines(text, fallback_firma=filename.rsplit(".", 1)[0])

        print("PARSED ROW COUNT:", len(rows))
        print("PARSED ROWS:", rows)

        return {
            "id": f"{filename}-image",
            "fileName": filename,
            "sourceType": "image",
            "label": f"Görsel - {filename}",
            "columns": [
                "urunKodu",
                "urunAciklamasi",
                "talepEdilenAdet",
                "firmaAdi",
                "firmaAdedi",
                "paraBirimi",
                "birimFiyat",
                "iskonto",
                "vade",
                "termin",
            ],
            "rows": rows,
        }

    return {
        "error": "Bu ilk sürümde sadece Excel, PDF ve görsel destekleniyor."
    }


# -----------------------------
# Mukayese / Öneri Endpoint
# -----------------------------
class CompareOffersRequest(BaseModel):
    rows: List[dict]
    exchangeRates: dict

def build_offer_analysis(rows, exchange_rates):
    comparison_rows = []
    grouped = {}

    for row in rows:
        urun_kodu = safe_str(row.get("urunKodu"))
        urun_aciklamasi = safe_str(row.get("urunAciklamasi"))
        firma_adi = safe_str(row.get("firmaAdi"))

        if not urun_aciklamasi and not urun_kodu:
            continue

        key = f"{urun_kodu.lower()}__{urun_aciklamasi.lower()}"

        normalized_para = safe_str(row.get("paraBirimi", "TRY")).upper() or "TRY"
        if normalized_para in ["₺", "TL"]:
            normalized_para = "TRY"

        normalized_row = {
            "urunKodu": urun_kodu,
            "urunAciklamasi": urun_aciklamasi,
            "talepEdilenAdet": safe_float(row.get("talepEdilenAdet", 0)),
            "firmaAdi": firma_adi,
            "firmaAdedi": safe_float(row.get("firmaAdedi", 0)),
            "paraBirimi": normalized_para,
            "birimFiyat": safe_float(row.get("birimFiyat", 0)),
            "iskonto": safe_float(row.get("iskonto", 0)),
            "netBirimFiyatTRY": get_net_birim_fiyat(row, exchange_rates),
            "vade": safe_str(row.get("vade", "")),
            "vadeGun": parse_vade_to_days(safe_str(row.get("vade", ""))),
            "termin": safe_str(row.get("termin", "")),
            "terminSkor": parse_termin_score(safe_str(row.get("termin", ""))),
        }

        comparison_rows.append(normalized_row)

        if key not in grouped:
            grouped[key] = []
        grouped[key].append(normalized_row)

    recommended_rows = []

    for _, teklifler in grouped.items():
        sorted_teklifler = sorted(
            teklifler,
            key=lambda x: (
                x["netBirimFiyatTRY"],
                -x["vadeGun"],
                x["terminSkor"],
                -x["firmaAdedi"],
            ),
        )

        best = sorted_teklifler[0]
        recommended_rows.append({
            "urunKodu": best["urunKodu"],
            "urunAciklamasi": best["urunAciklamasi"],
            "onerilenFirma": best["firmaAdi"],
            "onerilenParaBirimi": best["paraBirimi"],
            "onerilenBirimFiyat": best["birimFiyat"],
            "onerilenNetBirimFiyatTRY": best["netBirimFiyatTRY"],
            "firmaAdedi": best["firmaAdedi"],
            "vade": best["vade"],
            "termin": best["termin"],
        })

    return comparison_rows, recommended_rows


@app.post("/compare-offers")
def compare_offers(payload: CompareOffersRequest):
    rows = payload.rows
    exchange_rates = payload.exchangeRates or {"TRY": 1, "USD": 39.2, "EUR": 42.8}

    comparison_rows, recommended_rows = build_offer_analysis(rows, exchange_rates)

    return {
        "message": "Teklif karşılaştırması oluşturuldu.",
        "comparisonRows": comparison_rows,
        "recommendedRows": recommended_rows,
    }

@app.post("/export-comparison-report")
def export_comparison_report(payload: CompareOffersRequest):

    rows = payload.rows
    exchange_rates = payload.exchangeRates or {"TRY": 1, "USD": 39.2, "EUR": 42.8, "GBP": 41.2}

    comparison_rows, recommended_rows = build_offer_analysis(rows, exchange_rates)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Ozet"

    ws_comparison = wb.create_sheet("Mukayese")
    ws_recommended = wb.create_sheet("Oneri")

    bold_font = Font(bold=True, color="FFFFFF")
    dark_fill = PatternFill("solid", fgColor="1E3A8A")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # Özet sayfası
    ws_summary["A1"] = "TEKLIF KARŞILAŞTIRMA RAPORU"
    ws_summary["A1"].font = Font(bold=True, size=16)

    currencies = sorted(set([r["paraBirimi"] for r in comparison_rows])) if comparison_rows else []

    summary_rows = [
        ("Toplam Karşılaştırma Satırı", len(comparison_rows)),
        ("Önerilen Firma Satırı", len(recommended_rows)),
        ("Para Birimleri", ", ".join(currencies) if currencies else "-"),
        ("USD Kuru", exchange_rates.get("USD", "")),
        ("EUR Kuru", exchange_rates.get("EUR", "")),
        ("GBP Kuru", exchange_rates.get("GBP", "")),
    ]

    row_index = 3
    for label, value in summary_rows:
        ws_summary[f"A{row_index}"] = label
        ws_summary[f"B{row_index}"] = value
        ws_summary[f"A{row_index}"].font = Font(bold=True)
        row_index += 1

    # Mukayese sayfası
    comparison_headers = [
        "urunKodu",
        "urunAciklamasi",
        "firmaAdi",
        "firmaAdedi",
        "paraBirimi",
        "birimFiyat",
        "iskonto",
        "netBirimFiyatTRY",
        "vade",
        "termin",
    ]

    for col_index, header in enumerate(comparison_headers, start=1):
        cell = ws_comparison.cell(row=1, column=col_index, value=header)
        cell.font = bold_font
        cell.fill = dark_fill
        cell.alignment = center
        cell.border = border

    for row_index, item in enumerate(comparison_rows, start=2):
        for col_index, header in enumerate(comparison_headers, start=1):
            cell = ws_comparison.cell(row=row_index, column=col_index, value=item.get(header, ""))
            cell.border = border

    # Öneri sayfası
    recommended_headers = [
        "urunKodu",
        "urunAciklamasi",
        "onerilenFirma",
        "onerilenParaBirimi",
        "onerilenBirimFiyat",
        "onerilenNetBirimFiyatTRY",
        "firmaAdedi",
        "vade",
        "termin",
    ]

    for col_index, header in enumerate(recommended_headers, start=1):
        cell = ws_recommended.cell(row=1, column=col_index, value=header)
        cell.font = bold_font
        cell.fill = dark_fill
        cell.alignment = center
        cell.border = border

    for row_index, item in enumerate(recommended_rows, start=2):
        for col_index, header in enumerate(recommended_headers, start=1):
            cell = ws_recommended.cell(row=row_index, column=col_index, value=item.get(header, ""))
            cell.border = border

    # Kolon genişlikleri
    for ws in [ws_summary, ws_comparison, ws_recommended]:
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    value_length = len(str(cell.value)) if cell.value is not None else 0
                    if value_length > max_length:
                        max_length = value_length
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="teklif_karsilastirma_raporu.xlsx"'
        },
    )
# -----------------------------
# Merge Sources - eski yapı
# -----------------------------
class MergeRow(BaseModel):
    urunKodu: str | None = ""
    urunAciklamasi: str
    miktar: float | int | str
    birim: str


class MergeRowsRequest(BaseModel):
    rows: List[MergeRow]


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_quantity(value: Any) -> float:
    text = str(value or "").strip().replace(",", ".")
    filtered = ""

    for ch in text:
        if ch.isdigit() or ch in [".", "-"]:
            filtered += ch

    try:
        return float(filtered) if filtered else 0
    except:
        return 0


@app.post("/merge-sources")
def merge_sources(payload: MergeRowsRequest):
    merged = {}
    auto_code_map = {}
    auto_code_counter = 1
    match_candidates = []

    def generate_auto_code():
        nonlocal auto_code_counter
        code = f"PRD-{auto_code_counter:04}"
        auto_code_counter += 1
        return code

    def desc_key_of(aciklama: str, birim: str) -> str:
        return f"{aciklama.lower()}__{birim.lower()}"

    description_to_existing = {}
    for row in payload.rows:
        urun_kodu = normalize_text(row.urunKodu)
        urun_aciklamasi = normalize_text(row.urunAciklamasi)
        birim = normalize_text(row.birim)

        if urun_kodu and urun_aciklamasi:
            description_to_existing[desc_key_of(urun_aciklamasi, birim)] = {
                "urunKodu": urun_kodu,
                "urunAciklamasi": urun_aciklamasi,
            }

    for row in payload.rows:
        urun_kodu = normalize_text(row.urunKodu)
        urun_aciklamasi = normalize_text(row.urunAciklamasi)
        birim = normalize_text(row.birim)
        miktar = normalize_quantity(row.miktar)

        if not urun_aciklamasi:
            continue

        desc_key = desc_key_of(urun_aciklamasi, birim)

        if not urun_kodu:
            if desc_key in description_to_existing:
                suggested = description_to_existing[desc_key]

                urun_kodu = generate_auto_code()

                match_candidates.append({
                    "id": f"{urun_kodu}__{desc_key}",
                    "newCode": urun_kodu,
                    "newDescription": urun_aciklamasi,
                    "suggestedCode": suggested["urunKodu"],
                    "suggestedDescription": suggested["urunAciklamasi"],
                })
            else:
                if desc_key not in auto_code_map:
                    auto_code_map[desc_key] = generate_auto_code()
                urun_kodu = auto_code_map[desc_key]

        key = f"{urun_kodu.lower()}__{birim.lower()}"

        if key not in merged:
            merged[key] = {
                "urunKodu": urun_kodu,
                "urunAciklamasi": urun_aciklamasi,
                "miktar": miktar,
                "birim": birim,
            }
        else:
            merged[key]["miktar"] += miktar

    result = []
    for i, item in enumerate(merged.values(), start=1):
        result.append({
            "sira": i,
            "urunKodu": item["urunKodu"],
            "urunAciklamasi": item["urunAciklamasi"],
            "miktar": item["miktar"],
            "birim": item["birim"],
        })

    auto_assigned_count = sum(
        1 for item in result if item["urunKodu"].startswith("PRD-")
    )

    return {
        "message": "Kaynaklar birleştirildi.",
        "rows": result,
        "matchCandidates": match_candidates,
        "autoAssignedCount": auto_assigned_count,
    }