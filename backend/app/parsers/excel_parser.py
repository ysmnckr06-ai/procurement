import pandas as pd
import re
import os
import unicodedata
from openpyxl import load_workbook


def normalize_col(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))

    tr_map = str.maketrans({
        "ı": "i", "ğ": "g", "ü": "u",
        "ş": "s", "ö": "o", "ç": "c"
    })
    text = text.translate(tr_map)

    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def canonical_section_name(value):
    return ""


def money_equals(left, right, tolerance=0.05):
    return abs(float(left or 0) - float(right or 0)) <= tolerance


def row_numbers(values):
    numbers = []

    for value in values:
        number = clean_number(value)

        if number > 0:
            numbers.append(number)

    return numbers


def choose_kdv_excluded_total(numbers):
    positive = sorted({round(float(number), 2) for number in numbers if float(number or 0) > 0})

    if not positive:
        return 0

    if len(positive) >= 2:
        for lower in positive:
            for higher in positive:
                if higher <= lower:
                    continue

                if money_equals(higher, lower * 1.20, max(0.50, lower * 0.01)):
                    return lower

    return positive[-1]


def choose_section_total(row_total, numbers):
    net_total = choose_kdv_excluded_total(numbers)

    if row_total > 0 and net_total > 0 and money_equals(row_total, net_total * 1.20, max(0.50, net_total * 0.01)):
        return net_total

    return row_total or net_total


def is_meaningful_section_label(value):
    text = clean_text(value)

    if not text:
        return False

    norm = normalize_col(text)

    if not norm or norm in ["-", "_", "tl", "try", "eur", "usd"]:
        return False

    if re.fullmatch(r"[0-9\s.,/-]+", text):
        return False

    return bool(re.search(r"[a-z0-9]", norm))


def is_section_fill_rgb(rgb):
    if not rgb:
        return False

    value = str(rgb).replace("#", "").upper()

    if len(value) == 8:
        value = value[2:]

    if len(value) != 6:
        return False

    try:
        red = int(value[0:2], 16)
        green = int(value[2:4], 16)
        blue = int(value[4:6], 16)
    except ValueError:
        return False

    if red >= 245 and green >= 245 and blue >= 245:
        return False

    return red >= 200 and green >= 150 and blue <= 190


def highlighted_rows(file_path, sheet_name=None):
    try:
        workbook = load_workbook(file_path, read_only=False, data_only=True)
    except Exception:
        return set()

    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.worksheets[0]
    rows = set()

    for row in sheet.iter_rows():
        filled_cells = 0

        for cell in row:
            fill = cell.fill

            if not fill or fill.fill_type in [None, "none"]:
                continue

            fg_color = fill.fgColor
            rgb = fg_color.rgb if fg_color and fg_color.type == "rgb" else ""

            if (
                fill.fill_type == "solid"
                and fg_color
                and fg_color.type == "indexed"
                and fg_color.indexed in [5, 6, 13, 19, 22, 36, 44]
            ):
                filled_cells += 1
                continue

            if is_section_fill_rgb(rgb):
                filled_cells += 1

        if filled_cells > 0:
            rows.add(row[0].row)

    workbook.close()
    return rows


def section_name_from_cells(cells):
    text_cells = [
        clean_text(cell)
        for cell in cells
        if is_meaningful_section_label(cell)
    ]

    if not text_cells:
        return ""

    return clean_text(text_cells[0]).upper()


def has_material_identity(value):
    text = clean_text(value)

    if not text:
        return False

    if re.fullmatch(r"[0\s.,/-]+", text):
        return False

    return normalize_col(text) not in ["0", "00", "yok", "none", "nan"]


def looks_like_section_row(cells, code, brand, desc, qty, unit_price, net_price, row_total):
    if not desc or qty <= 0 or row_total <= 0:
        return False

    if should_skip_context_line(" ".join(clean_text(cell) for cell in cells)):
        return False

    has_code = has_material_identity(code)
    has_brand = has_material_identity(brand)
    has_unit_price = unit_price > 0 or net_price > 0

    if has_code or has_brand or has_unit_price:
        return False

    return bool(section_name_from_cells(cells))


def clean_text(val):
    if val is None or pd.isna(val):
        return ""
    text = str(val).strip()
    if text.lower() in ["nan", "none", "null"]:
        return ""
    return text


def clean_number(val):
    if val is None or pd.isna(val):
        return 0.0

    text = str(val).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))

    text = text.upper()
    text = text.replace("₺", "").replace("€", "").replace("£", "")
    text = text.replace("â‚º", "").replace("â‚¬", "").replace("Â£", "").replace("$", "")
    text = text.replace("TL", "").replace("TRY", "").replace("USD", "").replace("EUR", "").replace("GBP", "")
    text = text.replace("%", "")
    text = text.replace(",", ".")

    if text.count(".") > 1:
        parts = text.split(".")
        text = "".join(parts[:-1]) + "." + parts[-1]

    text = re.sub(r"[^0-9.\-]", "", text)

    try:
        return float(text)
    except Exception:
        return 0.0


def detect_currency_from_text(value):
    text = str(value or "").upper()

    if "$" in text or re.search(r"\b(USD|DOLAR)\b", text):
        return "USD"
    if "€" in text or "â‚¬" in text or "Ã¢â€šÂ¬" in text or re.search(r"\b(EUR|EURO)\b", text):
        return "EUR"
    if "£" in text or "Â£" in text or re.search(r"\b(GBP|STERLIN)\b", text):
        return "GBP"

    return "TRY"


def currency_from_token(value):
    text = normalize_col(value)

    if text in ["eur", "euro"]:
        return "EUR"
    if text in ["usd", "dolar", "dollar"]:
        return "USD"
    if text in ["tl", "try", "turk lirasi", "turk lira"]:
        return "TRY"
    if text in ["gbp", "sterlin"]:
        return "GBP"

    return ""


def detect_currency_token(value):
    text = str(value or "").upper()

    if "$" in text or re.search(r"\b(USD|DOLAR|DOLLAR)\b", text):
        return "USD"
    if "â‚¬" in text or "Ã¢â€šÂ¬" in text or "ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬" in text or re.search(r"\b(EUR|EURO)\b", text):
        return "EUR"
    if "Â£" in text or "Ã‚Â£" in text or re.search(r"\b(GBP|STERLIN)\b", text):
        return "GBP"
    if re.search(r"\b(TL|TRY)\b", text):
        return "TRY"

    return ""


def detect_excel_currency_context(raw_df):
    brand_currency = {}
    currency_counts = {}

    for _, row in raw_df.head(35).iterrows():
        for value in row.values:
            text = clean_text(value)

            if not text:
                continue

            detected = detect_currency_token(text)
            if detected:
                currency_counts[detected] = currency_counts.get(detected, 0) + 1

            for brand, currency in re.findall(r"([A-ZÇĞİÖŞÜ0-9 .&/_-]{2,35})\s*:\s*(EURO|EUR|DOLAR|USD|TL|TRY|GBP|STERLIN)\b", text.upper()):
                clean_brand = normalize_col(brand)
                clean_currency = currency_from_token(currency)

                if clean_brand and clean_currency and clean_brand not in ["tarih", "telefon", "faks", "email", "e mail", "teklif no"]:
                    brand_currency[clean_brand] = clean_currency

            plain_match = re.fullmatch(r"\s*([A-ZÇĞİÖŞÜ0-9 .&/_-]{2,35})\s+(EURO|EUR|DOLAR|USD|TL|TRY|GBP|STERLIN)\s*", text.upper())
            if plain_match:
                clean_brand = normalize_col(plain_match.group(1))
                clean_currency = currency_from_token(plain_match.group(2))

                if clean_brand and clean_currency and clean_brand not in ["euro", "dolar", "tl", "try"]:
                    brand_currency[clean_brand] = clean_currency

    non_try_counts = {currency: count for currency, count in currency_counts.items() if currency != "TRY"}
    default_currency = "TRY"

    if non_try_counts:
        default_currency = max(non_try_counts.items(), key=lambda item: item[1])[0]
    elif currency_counts:
        default_currency = max(currency_counts.items(), key=lambda item: item[1])[0]

    return brand_currency, default_currency


def find_header_row(df):
    strict_row = find_header_row_strict(df)
    if strict_row is not None:
        return strict_row

    return 0


def find_header_row_strict(df):
    for i in range(len(df)):
        row_text = " ".join(normalize_col(x) for x in df.iloc[i].values)

        has_desc = (
            "urun aciklamasi" in row_text
            or "aciklama" in row_text
            or "malzeme adi" in row_text
            or "malzeme tanimi" in row_text
            or "malzemenin cinsi" in row_text
        )
        has_qty = (
            "miktar" in row_text
            or "adet" in row_text
            or "quantity" in row_text
            or "ad mt" in row_text
        )
        has_price = (
            "birim fiyat" in row_text
            or "unit price" in row_text
            or "net birim fiyat" in row_text
            or "net tutar" in row_text
        )

        if has_desc and has_qty and has_price:
            return i

    return None


def read_best_excel_sheet(file_path):
    excel = pd.ExcelFile(file_path)
    best = None

    for sheet_name in excel.sheet_names:
        try:
            sample_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=160)
        except Exception:
            continue

        header_row = find_header_row_strict(sample_df)

        if header_row is None:
            continue

        nonempty_after_header = len(sample_df.iloc[header_row + 1:].dropna(how="all"))
        sheet_bonus = 1000 if "teklif" in normalize_col(sheet_name) else 0
        score = sheet_bonus + nonempty_after_header

        if best is None or score > best["score"]:
            best = {
                "sheet_name": sheet_name,
                "header_row": header_row,
                "score": score,
            }

    if best:
        return pd.read_excel(file_path, sheet_name=best["sheet_name"], header=None), best["sheet_name"]

    return pd.read_excel(file_path, header=None), excel.sheet_names[0] if excel.sheet_names else None


def find_col_exact_or_contains(columns, keywords, exclude_keywords=None):
    exclude_keywords = exclude_keywords or []
    normalized = {col: normalize_col(col) for col in columns}

    def allowed(clean):
        return not any(normalize_col(ex) in clean for ex in exclude_keywords)

    for col, clean in normalized.items():
        if not allowed(clean):
            continue
        for key in keywords:
            if normalize_col(key) == clean:
                return col

    for col, clean in normalized.items():
        if not allowed(clean):
            continue
        for key in keywords:
            if normalize_col(key) in clean:
                return col

    return None


def unique_columns(columns):
    seen = {}
    result = []

    for column in columns:
        clean = clean_text(column)

        if not clean:
            clean = "column"

        count = seen.get(clean, 0)
        seen[clean] = count + 1
        result.append(clean if count == 0 else f"{clean}__{count + 1}")

    return result


def should_skip_context_line(line):
    norm = normalize_col(line)

    if re.search(r"\b[tf]\s*:\s*\+?\s*\d", norm) or re.search(r"\b(tel|telefon|fax|faks)\b.*\d", norm):
        return True

    skip_words = [
        "toplam",
        "dip toplam",
        "ara toplam",
        "genel toplam",
        "kdv",
        "vade",
        "termin",
        "teslim",
        "firma",
        "teklif",
        "not",
        "rapor",
        "rapor olusturan",
        "rapor no",
        "mukayese",
        "karsilastirma raporu",
        "teklif karsilastirma",
        "telefon",
        "tel",
        "faks",
        "fax",
        "mail",
        "email",
        "www",
        "musteri onayi",
        "musteri onay",
        "teyit ediniz",
        "siparisin onayi",
        "siparis onayi",
        "teknik cizim",
        "uretime baslanacak",
        "teslimat suresi",
        "uretim hatalari",
        "saha revizyonlari",
        "servis ucreti",
        "konaklama",
        "ulasim ucreti",
    ]

    return any(word in norm for word in skip_words)


def detect_firma_adi(df, fallback_firma, file_name):
    header_aliases = {"firma", "firma adi", "firma adı", "tedarikci", "tedarikçi", "satici", "satıcı", "supplier"}
    for _, row in df.head(8).iterrows():
        cells = [clean_text(x) for x in row.values if clean_text(x)]
        if not cells:
            continue

        joined = " ".join(cells)
        norm = normalize_col(joined)

        if "firma" in norm or "tedarikci" in norm or "satici" in norm:
            for i, cell in enumerate(cells):
                c = normalize_col(cell)
                if "firma" in c or "tedarikci" in c or "satici" in c:
                    if c in header_aliases:
                        for below_index in range(1, min(len(df), 12)):
                            value = clean_text(df.iloc[below_index, i] if i < len(df.columns) else "")
                            value_norm = normalize_col(value)
                            if value and value_norm not in header_aliases and not should_skip_context_line(value):
                                return value
                    if i + 1 < len(cells):
                        candidate = cells[i + 1]
                        candidate_norm = normalize_col(candidate)
                        if candidate_norm not in header_aliases and not any(x in candidate_norm for x in ["urun kodu", "malzeme kodu", "aciklama", "miktar"]):
                            return candidate

        if len(cells) == 1:
            one = cells[0]
            one_norm = normalize_col(one)
            if (
                len(one) >= 3
                and "teklif" not in one_norm
                and "urun" not in one_norm
                and "miktar" not in one_norm
                and "fiyat" not in one_norm
            ):
                return one

    if fallback_firma:
        return fallback_firma

    return os.path.splitext(file_name)[0].replace("_", " ").replace("-", " ").title()


def detect_footer_info(df):
    vade = ""
    termin = ""
    dip_toplam = 0.0
    kdv = 0.0
    genel_toplam = 0.0
    context_text = []

    for _, row in df.iterrows():
        cells = [clean_text(x) for x in row.values]
        joined = " ".join(cells)
        norm = normalize_col(joined)
        nums = [clean_number(x) for x in cells if clean_number(x) > 0]
        context_text.append(joined)

        if "vade" in norm or "odeme" in norm:
            vade = joined

        if "termin" in norm or "teslim" in norm:
            termin = joined

        if "dip toplam" in norm or "ara toplam" in norm:
            if nums:
                dip_toplam = nums[-1]

        if "kdv" in norm:
            if nums:
                kdv = nums[-1]

        if "genel toplam" in norm or "yekun" in norm:
            if nums:
                genel_toplam = nums[-1]

    normalized_context = normalize_col(" ".join(context_text[:20]))

    if not vade:
        vade_match = re.search(r"(\d{1,3})\s*gun", normalized_context)
        if vade_match:
            vade = f"{vade_match.group(1)} gün"

    if not termin:
        termin_match = re.search(
            r"(\d{1,2}\s*-\s*\d{1,2}\s*hafta|\d{1,3}\s*gun|hazir|stok)",
            normalized_context,
        )
        if termin_match:
            termin = termin_match.group(1)

    return {
        "vade": vade,
        "termin": termin,
        "dipToplam": dip_toplam,
        "kdv": kdv,
        "genelToplam": genel_toplam,
    }


def parse_excel(file_path, firma_adi="", file_name=""):
    audit = parse_excel_with_audit(file_path, firma_adi, file_name)
    return audit["rows"]


def parse_excel_with_audit(file_path, firma_adi="", file_name=""):
    raw_df, selected_sheet_name = read_best_excel_sheet(file_path)

    firma = detect_firma_adi(raw_df, firma_adi, file_name)
    footer = detect_footer_info(raw_df)
    brand_currency_map, default_sheet_currency = detect_excel_currency_context(raw_df)
    header_row = find_header_row(raw_df)

    df = raw_df.copy()
    df.columns = unique_columns(df.iloc[header_row])
    df = df[header_row + 1:]
    df = df.dropna(how="all")

    code_col = find_col_exact_or_contains(df.columns, [
        "referans",
        "ref",
        "ref no",
        "urun kodu",
        "malzeme kodu",
        "stok kodu",
        "kod",
    ], exclude_keywords=["sira", "sıra", "s no"])

    desc_col = find_col_exact_or_contains(df.columns, [
        "malzemenin cinsi",
        "malzeme cinsi",
        "urun aciklamasi",
        "aciklama",
        "malzeme adi",
        "malzeme tanimi",
        "ürün",
        "urun",
        "ÜRÜN",
        "URUN",
    ], exclude_keywords=["kod", "sira", "sıra"])

    brand_col = find_col_exact_or_contains(df.columns, [
        "malzeme markasi",
        "urun markasi",
        "marka",
        "brand",
        "uretici",
        "üretici",
    ])

    qty_col = find_col_exact_or_contains(df.columns, [
        "ad/mt",
        "ad mt",
        "ad",
        "miktar",
        "adet",
        "teklif miktari",
        "firma adedi",
        "quantity",
    ])

    unit_col = find_col_exact_or_contains(df.columns, [
        "birim",
        "unit",
    ], exclude_keywords=["fiyat"])

    price_col = find_col_exact_or_contains(df.columns, [
        "birim fiyat",
        "unit price",
        "fiyat",
        "FIYAT",
        "FİYAT",
        "FIYAT",
    ], exclude_keywords=[
        "iskontolu",
        "net",
        "toplam",
        "tutar",
        "satir",
        "satır",
        "iskonto",
        "indirim",
    ])

    discount_col = find_col_exact_or_contains(df.columns, [
        "iskonto",
        "indirim",
        "discount",
    ], exclude_keywords=[
        "iskontolu fiyat",
        "net fiyat",
    ])

    net_price_col = find_col_exact_or_contains(df.columns, [
        "net birim fiyat tl",
        "net birim fiyat",
        "iskontolu fiyat",
        "net fiyat",
    ])

    total_col = find_col_exact_or_contains(df.columns, [
        "net tutar tl",
        "net tutar",
        "satir toplami",
        "satir toplam",
        "toplam tutar",
        "tutar",
        "toplam",
        "net toplam",
    ], exclude_keywords=[
        "genel toplam",
        "dip toplam",
        "ara toplam",
    ])

    currency_col = find_col_exact_or_contains(df.columns, [
        "para birimi",
        "doviz",
        "döviz",
        "currency",
    ])

    firma_col = find_col_exact_or_contains(df.columns, [
        "firma",
        "firma adi",
        "firma adÄ±",
        "tedarikci",
        "tedarikÃ§i",
        "satici",
        "satÄ±cÄ±",
        "supplier",
    ])

    delivery_col = find_col_exact_or_contains(df.columns, [
        "teslim",
        "termin",
        "teslim suresi",
        "lead time",
    ])

    payment_col = find_col_exact_or_contains(df.columns, [
        "vade",
        "odeme",
        "payment",
    ])

    rows = []
    sections = []
    errors = []
    warnings = []
    pending_rows = []

    def attach_section_to_pending(section):
        if not pending_rows:
            quantity = section.get("section_quantity") or 0
            section_total = section.get("section_total") or 0
            unit_price = section_total / quantity if quantity > 0 and section_total > 0 else 0
            rows.append({
                "firma": firma,
                "firmaAdi": firma,
                "urunKodu": section.get("product_code") or "",
                "urunAciklamasi": section["section_name"],
                "birim": section.get("unit") or "adet",
                "firmaAdedi": quantity,
                "paraBirimi": section.get("currency") or default_sheet_currency or "TRY",
                "birimFiyat": unit_price,
                "iskonto": 0,
                "netBirimFiyat": unit_price,
                "netToplam": section_total,
                "netBirimFiyatDosyadan": unit_price,
                "satirToplamDosyadan": section_total,
                "vade": footer.get("vade", ""),
                "termin": footer.get("termin", ""),
                "firmaDipToplam": footer.get("dipToplam", 0),
                "firmaKdv": footer.get("kdv", 0),
                "firmaGenelToplam": footer.get("genelToplam", 0),
                "kaynakDosya": file_name,
                "kaynakTipi": "excel",
                "parserUyarilari": [],
                "section_name": "",
                "section_total": 0,
                "section_quantity": 0,
                "price_status": "flat_main_item",
            })
            warnings.append("Bu dosyada alt kalem ilişkisi bulunmayan ana kalem satırı görüldü; bağımsız ana kalem olarak aktarılabilir.")
            return

        for row in pending_rows:
            row["section_name"] = section["section_name"]
            row["section_total"] = section["section_total"]
            row["section_quantity"] = section.get("section_quantity") or 0
            row["birimFiyat"] = 0
            row["netBirimFiyat"] = 0
            row["netToplam"] = 0
            row["netBirimFiyatDosyadan"] = 0
            row["satirToplamDosyadan"] = 0
            row["price_status"] = "section_total_only"

        warnings.append(
            f"{section['section_name']} toplamı üstündeki {len(pending_rows)} malzemeye bağlandı; parça fiyatları aktarılmadı."
        )
        pending_rows.clear()

    for row_index, r in df.iterrows():
        row_firma = clean_text(r.get(firma_col)) if firma_col is not None else firma
        if not row_firma or normalize_col(row_firma) in ["firma", "firma adi", "firma adÄ±", "tedarikci", "tedarikÃ§i", "satici", "satÄ±cÄ±", "supplier"]:
            row_firma = firma
        code = clean_text(r.get(code_col)) if code_col is not None else ""
        brand = clean_text(r.get(brand_col)) if brand_col is not None else ""
        desc = clean_text(r.get(desc_col)) if desc_col is not None else ""
        cells = [clean_text(x) for x in r.values]
        joined = " ".join(cells)

        desc_norm = normalize_col(desc)
        numbers = row_numbers(cells)
        row_total_from_file = clean_number(r.get(total_col)) if total_col is not None else 0.0
        if row_total_from_file <= 0 and numbers:
            row_total_from_file = numbers[-1]
        qty = clean_number(r.get(qty_col)) if qty_col is not None else 0.0

        if not code and not desc:
            continue

        if should_skip_context_line(joined):
            continue

        unit = clean_text(r.get(unit_col)) if unit_col is not None else "adet"

        price = clean_number(r.get(price_col)) if price_col is not None else 0.0
        discount = clean_number(r.get(discount_col)) if discount_col is not None else 0.0
        net_price_from_file = clean_number(r.get(net_price_col)) if net_price_col is not None else 0.0
        explicit_currency = detect_currency_token(r.get(currency_col)) if currency_col is not None else detect_currency_token(joined)
        brand_currency = brand_currency_map.get(normalize_col(brand), "")
        row_currency = explicit_currency or brand_currency or default_sheet_currency or "TRY"
        row_vade = clean_text(r.get(payment_col)) if payment_col is not None else footer.get("vade", "")
        row_termin = clean_text(r.get(delivery_col)) if delivery_col is not None else footer.get("termin", "")
        row_warnings = []

        if looks_like_section_row(cells, code, brand, desc, qty, price, net_price_from_file, row_total_from_file):
            section_total = choose_section_total(row_total_from_file, numbers)
            display_section_name = section_name_from_cells(cells) or f"BÖLÜM-{len(sections) + 1}"
            section = {
                "section_name": display_section_name,
                "section_total": section_total,
                "section_quantity": qty if qty > 0 else 0,
                "product_code": code,
                "unit": unit or "adet",
                "currency": row_currency,
            }
            if qty <= 0:
                warnings.append(f"{display_section_name} ana kalem miktarı kontrol gerekli.")
            sections.append(section)
            attach_section_to_pending(section)
            continue

        if price <= 0 and net_price_from_file > 0:
            price = net_price_from_file

        if discount <= 0 and price > 0 and net_price_from_file > 0 and net_price_from_file < price:
            discount = round((1 - (net_price_from_file / price)) * 100, 4)

        net_unit_calculated = price * (1 - discount / 100)
        row_total_calculated = net_unit_calculated * qty if qty > 0 else 0.0

        if row_total_calculated > 0 and row_total_from_file <= 0:
            row_total_from_file = row_total_calculated
        elif row_total_calculated > 0 and row_total_from_file > 0 and not money_equals(row_total_calculated, row_total_from_file, 0.10):
            row_warnings.append(
                f"Satır toplamı kontrol edilmeli: {desc} ({qty} x {net_unit_calculated:.2f} != {row_total_from_file:.2f})"
            )

        if not desc:
            continue

        if qty <= 0:
            continue

        price_status = "line_priced"

        if price <= 0 or row_total_from_file <= 0:
            price = 0
            net_price_from_file = 0
            row_total_from_file = 0
            price_status = "pending_section_total"

        if price_status == "line_priced" and row_total_from_file > 0 and qty > 0 and not money_equals(row_total_calculated, row_total_from_file, 0.10) and net_price_from_file <= 0:
            price = row_total_from_file / qty
            net_price_from_file = price
            discount = 0
            net_unit_calculated = price

        net_unit_value = net_price_from_file or net_unit_calculated or price

        rows.append({
            "firma": row_firma,
            "firmaAdi": row_firma,
            "urunKodu": code,
            "marka": brand,
            "brand": brand,
            "urunAciklamasi": desc,
            "birim": unit or "adet",
            "firmaAdedi": qty,
            "paraBirimi": row_currency,
            "birimFiyat": price,
            "iskonto": discount,
            "netBirimFiyat": net_unit_value,
            "netToplam": row_total_from_file,
            "netBirimFiyatDosyadan": net_price_from_file,
            "satirToplamDosyadan": row_total_from_file,
            "vade": row_vade,
            "termin": row_termin,
            "firmaDipToplam": footer.get("dipToplam", 0),
            "firmaKdv": footer.get("kdv", 0),
            "firmaGenelToplam": footer.get("genelToplam", 0),
            "kaynakDosya": file_name,
            "kaynakTipi": "excel",
            "parserUyarilari": row_warnings,
            "section_name": "",
            "section_total": 0,
            "section_quantity": 0,
            "price_status": price_status,
        })
        pending_rows.append(rows[-1])

    if sections and pending_rows:
        for row in pending_rows:
            row["birimFiyat"] = 0
            row["netBirimFiyat"] = 0
            row["netToplam"] = 0
            row["netBirimFiyatDosyadan"] = 0
            row["satirToplamDosyadan"] = 0
            row["price_status"] = "section_total_only"

        warnings.append(
            f"{len(pending_rows)} Excel malzeme satırı için sonraki kategori toplamı bulunamadı; parça fiyatları güvenli olarak aktarılmadı."
        )
        pending_rows.clear()

    ungrouped_pending = [row for row in pending_rows if row.get("price_status") == "pending_section_total"]
    if ungrouped_pending:
        warnings.append(f"{len(ungrouped_pending)} Excel malzeme satırında fiyat yok; satırlar 0 fiyatla aktarıldı.")

    checked_total = footer.get("dipToplam") or footer.get("genelToplam") or 0
    product_total = sum(float(row.get("satirToplamDosyadan") or 0) for row in rows)

    if checked_total > 0 and product_total > 0 and not money_equals(product_total, checked_total, 0.50):
        errors.append(
            f"Excel genel toplam tutmuyor: ürünler {product_total:.2f}, teklif {checked_total:.2f}"
        )

    return {
        "rows": rows,
        "sections": sections,
        "errors": errors,
        "warnings": warnings,
    }
